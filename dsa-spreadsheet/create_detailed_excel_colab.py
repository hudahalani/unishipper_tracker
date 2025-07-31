import csv
import os
from datetime import datetime
import re
from io import StringIO
import pandas as pd
from google.colab import files
import zipfile

def create_detailed_excel_colab():
    # Upload CSV file
    print("Please upload your CSV file...")
    uploaded = files.upload()
    
    if not uploaded:
        print("No file uploaded!")
        return
    
    # Get the uploaded file
    csv_filename = list(uploaded.keys())[0]
    csv_content = uploaded[csv_filename]
    
    print(f"Processing file: {csv_filename}")
    
    # Extract date from filename
    date_match = re.search(r'(\d{2}_\d{2}_\d{4})', csv_filename)
    if not date_match:
        print("Could not extract date from filename")
        return
    
    filename_date = date_match.group(1)
    print(f"Date from filename: {filename_date}")
    
    # Store all records
    all_records = []
    
    # Parse CSV from the uploaded content
    try:
        # Try to decode as UTF-8 first
        csv_text = csv_content.decode('utf-8')
    except UnicodeDecodeError:
        # If UTF-8 fails, try other encodings
        for encoding in ['latin-1', 'cp1252', 'iso-8859-1']:
            try:
                csv_text = csv_content.decode(encoding)
                print(f"Successfully decoded with {encoding} encoding")
                break
            except UnicodeDecodeError:
                continue
        else:
            print("Could not decode file with any encoding")
            return
    
    # Parse CSV from the content
    reader = csv.DictReader(StringIO(csv_text))
    
    for row in reader:
        # Check if created date matches filename date
        created_date_str = row.get('Created Date', '').strip()
        
        # Skip if status is VOID
        status = row.get('Status', '').strip()
        if status.upper() == 'VOID':
            continue
        
        # Parse created date and check if it matches filename date
        try:
            created_datetime = datetime.strptime(created_date_str, '%Y-%m-%d %H:%M:%S')
            created_date_only = created_datetime.strftime('%m_%d_%Y')
            
            # Only process records where created date matches filename date
            if created_date_only != filename_date:
                continue
                
        except ValueError as e:
            print(f"Error parsing date '{created_date_str}': {e}")
            continue
        
        # Extract PO number from Reference 1 (remove "PO Number - " prefix)
        po_reference = row.get('Reference 1', '').strip()
        po_number = po_reference.replace('PO Number - ', '') if po_reference.startswith('PO Number - ') else po_reference
        
        # Clean up pickup date (remove time, keep only date)
        pickup_date = row.get('Ship Date', '').strip()
        if pickup_date:
            try:
                # Parse the date and format it as YYYY-MM-DD
                pickup_datetime = datetime.strptime(pickup_date, '%Y-%m-%d %H:%M:%S')
                pickup_date_clean = pickup_datetime.strftime('%Y-%m-%d')
            except:
                pickup_date_clean = pickup_date
        else:
            pickup_date_clean = ''
        
        # Clean up shipper name (replace "Display Source Alliance" with "DSA")
        shipper_name = row.get('Sender Company Name', '').strip()
        shipper_clean = shipper_name.replace('Display Source Alliance', 'DSA').replace('DISPLAY SOURCE ALLIANCE', 'DSA')
        
        # Clean up weight (remove "lb" and keep only the number)
        weight_raw = row.get('Weight', '').strip()
        weight_clean = ''
        if weight_raw:
            # Extract just the number from "235 lb" -> "235"
            weight_match = re.search(r'(\d+)', weight_raw)
            if weight_match:
                weight_clean = weight_match.group(1)
        
        # Clean up carrier name (replace "SOUTHEASTERN FREIGHT LINES" with "SEFL", "XPO Logistics" with "XPO", and "RL Carriers" with "R&L")
        carrier_raw = row.get('Carrier', '').strip()
        carrier_clean = carrier_raw.replace('SOUTHEASTERN FREIGHT LINES', 'SEFL').replace('XPO Logistics', 'XPO').replace('RL Carriers', 'R&L')
        
        # Clean up status (replace "IN_TRANSIT" with "IN TRANSIT")
        status_raw = row.get('Status', '').strip()
        status_clean = status_raw.replace('IN_TRANSIT', 'IN TRANSIT')
        
        # Map CSV columns to requested Excel columns
        record = {
            'DSA PO#': po_number,
            'DSA Sales Order #': '',  # Not available in CSV
            'P/U Date': pickup_date_clean,
            'Shipper': shipper_clean,
            'Shipper Address': f"{row.get('Sender Address Line1', '')} {row.get('Sender Address Line2', '')}".strip(),
            'City': row.get('Sender City', '').strip(),
            'State': row.get('Sender State', '').strip(),
            'Zip': row.get('Sender Zip', '').strip(),
            'Consignee': row.get('Destination Company Name', '').strip(),
            'Store #': '',  # Extract store number from consignee name if possible
            'Consignee Address': f"{row.get('Receiver Address Line1', '')} {row.get('Receiver Address Line2', '')}".strip(),
            'City': row.get('Receiver City', '').strip(),
            'State': row.get('Receiver State', '').strip(),
            'Zip': row.get('Receiver Zip', '').strip(),
            'Pcs': row.get('Unit Count', '').strip(),
            'Wgt': '',  # Will be set below after cleaning
            'Dims 1': '',  # Not available in CSV
            'Dims 2': '',  # Not available in CSV
            'Dims 3': '',  # Not available in CSV
            'Dims 4': '',  # Not available in CSV
            'Carrier': carrier_clean,
            'Unishippers BOL#': row.get('BOL #', '').strip(),
            'Pro#': row.get('PRO/Tracking#', '').strip(),
            'Status': status_clean,
            'Est/Actual Delv Date': row.get('Estimated Delivery Date', '').strip(),
            'Install Date': '',  # Not available in CSV
            'Rate': row.get('Quoted Amount', '').strip()
        }
        
        # Set the cleaned weight value
        record['Wgt'] = weight_clean
        
        # Try to extract store number from consignee name
        consignee_name = row.get('Destination Company Name', '')
        store_match = re.search(r'#(\d+)', consignee_name)
        if store_match:
            record['Store #'] = store_match.group(1)
        
        all_records.append(record)
        print(f"Match found: BOL {record['Unishippers BOL#']}, PRO {record['Pro#']} - {record['Consignee']}")
    
    # Write results to Excel file
    output_filename = f"detailed_results_{filename_date}_by_store.xlsx"
    
    if all_records:
        # Group records by store type - 11 specific tabs
        store_groups = {
            'Albertsons': [],
            'Kroger': [],
            'HEB': [],
            'Ralphs': [],
            'WinCo': [],
            'Food Lion': [],
            'Hy-Vee': [],
            'Stop & Shop': [],
            'Meijer': [],
            'Weigel Stores': [],
            'Misc': []  # For any unmatched consignees
        }
        
        for record in all_records:
            consignee = record['Consignee']
            assigned = False
            
            # Check for specific store types
            if 'ALBERTSONS' in consignee.upper():
                store_groups['Albertsons'].append(record)
                assigned = True
            elif 'KROGER' in consignee.upper():
                store_groups['Kroger'].append(record)
                assigned = True
            elif 'HEB' in consignee.upper():
                store_groups['HEB'].append(record)
                assigned = True
            elif 'RALPHS' in consignee.upper():
                store_groups['Ralphs'].append(record)
                assigned = True
            elif 'WINCO' in consignee.upper():
                store_groups['WinCo'].append(record)
                assigned = True
            elif 'FOOD LION' in consignee.upper():
                store_groups['Food Lion'].append(record)
                assigned = True
            elif 'HY-VEE' in consignee.upper():
                store_groups['Hy-Vee'].append(record)
                assigned = True
            elif 'STOP & SHOP' in consignee.upper() or 'STOP AND SHOP' in consignee.upper():
                store_groups['Stop & Shop'].append(record)
                assigned = True
            elif 'MEIJER' in consignee.upper():
                store_groups['Meijer'].append(record)
                assigned = True
            elif 'WEIGEL' in consignee.upper():
                store_groups['Weigel Stores'].append(record)
                assigned = True
            
            # If no match found, put in Misc
            if not assigned:
                store_groups['Misc'].append(record)
        
        # Write to Excel with multiple sheets
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            # Create a summary sheet with all records
            df_all = pd.DataFrame(all_records)
            df_all.to_excel(writer, sheet_name='All Records', index=False)
            
            # Create separate sheets for each store type
            for store_type, records in store_groups.items():
                if records:  # Only create sheet if there are records
                    df_store = pd.DataFrame(records)
                    
                    # Clean sheet name (remove spaces and special characters)
                    sheet_name = store_type.replace(' ', '_').replace('&', 'and')
                    if len(sheet_name) > 31:  # Excel sheet name limit
                        sheet_name = sheet_name[:31]
                    
                    df_store.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get the workbook for formatting
            workbook = writer.book
            
            # Apply formatting to all sheets
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add header formatting
                from openpyxl.styles import Font, PatternFill
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color="FFFFFF")
        
        print(f"\nResults written to: {output_filename}")
        print(f"Total records processed: {len(all_records)}")
        print(f"Created sheets for: {', '.join(store_groups.keys())}")
        
        # Download the Excel file
        files.download(output_filename)
        
    else:
        # Create empty DataFrame with headers
        columns = ['DSA PO#', 'DSA Sales Order #', 'P/U Date', 'Shipper', 'Shipper Address', 'City', 'State', 'Zip', 
                  'Consignee', 'Store #', 'Consignee Address', 'City', 'State', 'Zip', 'Pcs', 'Wgt', 
                  'Dims 1', 'Dims 2', 'Dims 3', 'Dims 4', 'Carrier', 'Unishippers BOL#', 
                  'Pro#', 'Status', 'Est/Actual Delv Date', 'Install Date', 'Rate']
        df = pd.DataFrame(columns=columns)
        df.to_excel(output_filename, sheet_name='All Records', index=False)
        print(f"\nNo records found. Empty Excel file created: {output_filename}")
        files.download(output_filename)

# Run the function
if __name__ == "__main__":
    create_detailed_excel_colab() 